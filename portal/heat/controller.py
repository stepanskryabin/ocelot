from openpyxl import load_workbook
from decimal import getcontext, Decimal
from datetime import datetime, date
from typing import List, Type


# Дата снятия показаний ОДПУ и ИПУ, type int
READING_DATE = 18

# Актуальное время
CURRENT = datetime.now()

# Норматив на подогрев ГВС
NORM_GVS = '0.0718'


def parsing_xlsx(input_file: str,
                 start_row: int,
                 stop_row: int,
                 start_col: int,
                 stop_col: int,
                 sheet: str = None,
                 cells: tuple = None) -> List[list]:
    """Парсинг XLSX файлов. В результате будет список со вложенными
    списками содержащими данные из ячеек.

    Args:
        input_file (str): прямой или относительный путь к файлу
        start_row (int): номер начальной строки
        stop_row (int): номер завершающей строки
        start_col (int): номер начального столбца
        stop_col (int): номер заверщшающего столбца
        sheet (str, optional): имя листа. если оставить None,
        будет выбран первый лист в книге.
        cells (tuple, optional): ячейки которые необходимо спарсить.
        если оставить None, в результирующем списке будут все ячейки
        от start_col до stop_col

    Returns:
        List[list]: Списко со вложенным списком содержащим значения ячеек.
    """
    xlsx_file = load_workbook(filename=input_file, data_only=True)
    if sheet is None:
        sheet = xlsx_file[xlsx_file.sheetnames[0]]
    else:
        sheet = xlsx_file[sheet]
    result = []
    for row in range(start_row, stop_row, 1):
        list_cell = sheet.iter_rows(min_col=start_col, max_col=stop_col,
                                    min_row=row, max_row=row,
                                    values_only=True)
        new_list_row = []
        for cell in list_cell.__next__():
            if cell is None:
                new_list_row.append(0)
            else:
                new_list_row.append(cell)
        if cells is not None:
            new_list_row = [new_list_row[x] for x in cells]
        result.append(new_list_row)
    return result


def calc_finance(normativ: Type[Decimal],
                 tarif: Type[Decimal],
                 nalog: Type[Decimal]) -> Type[Decimal]:
    """Вычисление стоимости еденицы объёма

    Args:
        normativ (Decimal): [description]
        tarif (Decimal): [description]
        nalog (Decimal): [description]

    Returns:
        Decimal: [description]
    """
    return tarif * (normativ + (normativ * (nalog / Decimal('100'))))


def previous_date(date_time: int) -> Type[date]:
    """Дата предыдущего месяца

    Args:
        date_time (int): текущая дата
    """
    prev_month = date_time.month - 1
    prev_year = date_time.year
    if prev_month == 0:
        prev_month = 12
        prev_year = date_time.year - 1
    return date(prev_year,
                prev_month,
                READING_DATE)


def current_date() -> Type[date]:
    """Текущая дата, с фиксированным днём

    Returns:
        [type]: [description]
    """
    return date(CURRENT.year,
                CURRENT.month,
                READING_DATE)


def calc_calories(mass: Type[Decimal]) -> Type[Decimal]:
    """Вычисление объёма тепловой энергии в Гигакалориях
    из массы горячей воды

    Args:
        mass (Type[Decimal]): масса горячей воды

    Returns:
        Type[Decimal]: объём тепловой энергии, в Гкал.
    """
    getcontext().prec = 5
    return Decimal(mass) * Decimal(NORM_GVS)
